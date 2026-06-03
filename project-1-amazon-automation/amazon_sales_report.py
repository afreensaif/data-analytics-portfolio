import pandas as pd
import numpy as np
import xlwings as xw  
from datetime import datetime

input_file_path = r"C:\Users\afree\OneDrive\Desktop\Afreen Saif - Portfolio\Kaggle dataset - ecommerce sales\Amazon Sale Report.xlsx"
output_file = r"C:\Users\afree\OneDrive\Desktop\Afreen Saif - Portfolio\Kaggle dataset - ecommerce sales\Amazon Sale KPI Report.xlsx"


from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
import warnings
warnings.filterwarnings("ignore")


df = pd.read_excel(input_file_path,sheet_name=f'Amazon Sale Report')
df.columns = df.columns.str.strip()

df["Date"] = pd.to_datetime(df["Date"], format="%m-%d-%y", errors="coerce")
df["Month"]      = df["Date"].dt.to_period("M").astype(str)
df["Month_Name"] = df["Date"].dt.strftime("%b %Y")
 
# Clean Amount
df["Amount"] = pd.to_numeric(df["Amount"], errors="coerce").fillna(0)
df["Qty"]    = pd.to_numeric(df["Qty"],    errors="coerce").fillna(0)
 
# Classify order status
df["Is_Delivered"]  = df["Status"].str.contains("Delivered", case=False, na=False)
df["Is_Cancelled"]  = df["Status"].str.contains("Cancelled", case=False, na=False)
df["Is_Returned"]   = df["Status"].str.contains("Return",    case=False, na=False)
df["Is_Successful"] = df["Is_Delivered"]
 
delivered = df[df["Is_Delivered"]].copy()
print(f"   ✅ {len(df):,} orders loaded | {df['Date'].min().date()} → {df['Date'].max().date()}")

# ── STEP 2: CALCULATE KPIs ──────────────────────────────────
print("⏳ Step 2: Calculating KPIs...")
 
total_orders      = len(df)
total_revenue     = delivered["Amount"].sum()
total_units       = delivered["Qty"].sum()
avg_order_value   = delivered["Amount"].mean()
cancellation_rate = df["Is_Cancelled"].sum() / total_orders * 100
return_rate       = df["Is_Returned"].sum() / total_orders * 100
delivery_rate     = df["Is_Delivered"].sum() / total_orders * 100
b2b_revenue       = delivered[delivered["B2B"] == True]["Amount"].sum()
b2c_revenue       = delivered[delivered["B2B"] == False]["Amount"].sum()
 

summary_kpis = {
    "Total Orders":        f"{total_orders:,}",
    "Delivered Orders":    f"{df['Is_Delivered'].sum():,}",
    "Total Revenue (INR)": f"₹{total_revenue:,.0f}",
    "Total Units Sold":    f"{int(total_units):,}",
    "Avg Order Value":     f"₹{avg_order_value:,.0f}",
    "Delivery Rate":       f"{delivery_rate:.1f}%",
    "Cancellation Rate":   f"{cancellation_rate:.1f}%",
    "Return Rate":         f"{return_rate:.1f}%",
    "B2B Revenue":         f"₹{b2b_revenue:,.0f}",
    "B2C Revenue":         f"₹{b2c_revenue:,.0f}",
}

print("\n📊 Summary KPIs:" + "\n".join(f"   {k}: {v}" for k, v in summary_kpis.items()))

# print(df)
monthly = (
    delivered.groupby(["Month", "Month_Name"])
    .agg(Revenue=("Amount", "sum"), Orders=("Order ID", "count"), Units=("Qty", "sum"))
    .reset_index()
    .sort_values("Month")
)
monthly["AOV"] = (monthly["Revenue"] / monthly["Orders"]).round(0)
monthly["MoM_Growth"] = monthly["Revenue"].pct_change() * 100
 
# Category performance
category = (
    delivered.groupby("Category")
    .agg(Revenue=("Amount", "sum"), Orders=("Order ID", "count"), Units=("Qty", "sum"))
    .reset_index()
    .sort_values("Revenue", ascending=False)
)
category["Revenue_Share"] = (category["Revenue"] / category["Revenue"].sum() * 100).round(1)
category["AOV"] = (category["Revenue"] / category["Orders"]).round(0)
 


# Top 10 states by revenue
state = (
    delivered.groupby("ship-state")
    .agg(Revenue=("Amount", "sum"), Orders=("Order ID", "count"))
    .reset_index()
    .sort_values("Revenue", ascending=False)
    .head(10)
    .rename(columns={"ship-state": "State"})
)
 
# Fulfilment breakdown
fulfilment = (
    delivered.groupby("Fulfilment")
    .agg(Revenue=("Amount", "sum"), Orders=("Order ID", "count"))
    .reset_index()
    .sort_values("Revenue", ascending=False)
)
fulfilment["Revenue_Share"] = (fulfilment["Revenue"] / fulfilment["Revenue"].sum() * 100).round(1)
 
# Cancellation by category
cancel_by_cat = (
    df.groupby("Category")
    .agg(Total=("Order ID", "count"), Cancelled=("Is_Cancelled", "sum"))
    .reset_index()
)
cancel_by_cat["Cancel_Rate"] = (cancel_by_cat["Cancelled"] / cancel_by_cat["Total"] * 100).round(1)
cancel_by_cat = cancel_by_cat.sort_values("Cancel_Rate", ascending=False)
 
print("   ✅ KPIs calculated")

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
 
    # Sheet 1 — Summary KPIs
    summary_df = pd.DataFrame(list(summary_kpis.items()), columns=["KPI", "Value"])
    summary_df.to_excel(writer, sheet_name="📊 Executive Summary", index=False, startrow=3)

# Sheet 2 — Monthly Trend
    monthly_out = monthly[["Month_Name", "Revenue", "Orders", "Units", "AOV", "MoM_Growth"]].copy()
    monthly_out.columns = ["Month", "Revenue (INR)", "Orders", "Units Sold", "AOV (INR)", "MoM Growth %"]
    monthly_out.to_excel(writer, sheet_name="📈 Monthly Trend", index=False, startrow=2)
 
    # Sheet 3 — Category Performance
    cat_out = category.copy()
    cat_out.columns = ["Category", "Revenue (INR)", "Orders", "Units", "Revenue Share %", "AOV (INR)"]
    cat_out.to_excel(writer, sheet_name="🏷️ Category Performance", index=False, startrow=2)
 
    # Sheet 4 — Top States
    state.columns = ["State", "Revenue (INR)", "Orders"]
    state.to_excel(writer, sheet_name="📍 Top States", index=False, startrow=2)
 
    # Sheet 5 — Fulfilment
    fulfilment.columns = ["Fulfilment Type", "Revenue (INR)", "Orders", "Revenue Share %"]
    fulfilment.to_excel(writer, sheet_name="🚚 Fulfilment", index=False, startrow=2)
 
    # Sheet 6 — Cancellation Analysis
    cancel_out = cancel_by_cat.copy()
    cancel_out.columns = ["Category", "Total Orders", "Cancelled", "Cancel Rate %"]
    cancel_out.to_excel(writer, sheet_name="❌ Cancellations", index=False, startrow=2)
 
# Brand colors
DARK_TEAL   = "1A6B72"
LIGHT_TEAL  = "D6ECEE"
WHITE       = "FFFFFF"
DARK_GRAY   = "2D2D2D"
MID_GRAY    = "F2F2F2"
ACCENT      = "E8F5E9"


 
 
# ── STEP 4: STYLE THE WORKBOOK ──────────────────────────────
print("⏳ Step 4: Applying formatting...")
 
wb = load_workbook(output_file)
 
def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)
 
def make_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)
 
def style_header_row(ws, row, col_start, col_end, title=None):
    """Dark teal header row"""
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill      = make_fill(DARK_TEAL)
        cell.font      = Font(bold=True, color=WHITE, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = make_border()
    if title:
        ws.cell(row=row - 1, column=col_start).value = title
        ws.cell(row=row - 1, column=col_start).font  = Font(bold=True, size=12, color=DARK_GRAY)
 
def style_data_rows(ws, start_row, end_row, col_start, col_end):
    for r in range(start_row, end_row + 1):
        fill = make_fill(MID_GRAY) if r % 2 == 0 else make_fill(WHITE)
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill      = fill
            cell.border    = make_border()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font      = Font(size=10, color=DARK_GRAY)
 
def auto_width(ws, min_w=12, max_w=30):
    for col in ws.columns:
        length = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)
 
 
# ── SHEET 1: Executive Summary ───────────────────────────────
ws1 = wb["📊 Executive Summary"]
ws1.sheet_view.showGridLines = False
 
# Title banner
ws1.merge_cells("A1:B2")
ws1["A1"].value     = "🛒  Amazon Sales — Executive KPI Report"
ws1["A1"].font      = Font(bold=True, size=14, color=WHITE)
ws1["A1"].fill      = make_fill(DARK_TEAL)
ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
 
# Style header (row 4 after startrow=3)
style_header_row(ws1, row=4, col_start=1, col_end=2)
style_data_rows(ws1, start_row=5, end_row=5+len(summary_kpis)-1, col_start=1, col_end=2)
 
# Make KPI column left-aligned, value column right-aligned + bold
for r in range(5, 5 + len(summary_kpis)):
    ws1.cell(row=r, column=1).alignment = Alignment(horizontal="left",  indent=1)
    ws1.cell(row=r, column=2).alignment = Alignment(horizontal="right", indent=1)
    ws1.cell(row=r, column=2).font      = Font(bold=True, size=10, color=DARK_TEAL)
 
ws1.column_dimensions["A"].width = 26
ws1.column_dimensions["B"].width = 22
ws1.row_dimensions[1].height     = 36
ws1.row_dimensions[4].height     = 22
 
 
# ── SHEETS 2-6: Generic styling ─────────────────────────────
sheet_configs = {
    "📈 Monthly Trend":       {"title": "Monthly Revenue Trend",        "cols": 6, "data_start": 4},
    "🏷️ Category Performance": {"title": "Category Performance Breakdown","cols": 6, "data_start": 4},
    "📍 Top States":          {"title": "Top 10 States by Revenue",     "cols": 3, "data_start": 4},
    "🚚 Fulfilment":          {"title": "Fulfilment Type Breakdown",    "cols": 4, "data_start": 4},
    "❌ Cancellations":       {"title": "Cancellation Rate by Category", "cols": 4, "data_start": 4},
}
 
for sheet_name, cfg in sheet_configs.items():
    ws = wb[sheet_name]
    ws.sheet_view.showGridLines = False
 
    # Section title in row 1
    ws.merge_cells(f"A1:{get_column_letter(cfg['cols'])}1")
    ws["A1"].value     = cfg["title"]
    ws["A1"].font      = Font(bold=True, size=13, color=WHITE)
    ws["A1"].fill      = make_fill(DARK_TEAL)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
 
    # Header row (row 3, since startrow=2 puts headers at row 3)
    style_header_row(ws, row=3, col_start=1, col_end=cfg["cols"])
 
    # Data rows
    max_row = ws.max_row
    style_data_rows(ws, start_row=4, end_row=max_row, col_start=1, col_end=cfg["cols"])
    auto_width(ws)
 
    # Left-align first column (names/labels)
    for r in range(4, max_row + 1):
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", indent=1)
 
 
# ── ADD CHART: Monthly Revenue Bar Chart ────────────────────
ws2 = wb["📈 Monthly Trend"]
n_months = len(monthly)
 
chart = BarChart()
chart.type    = "col"
chart.title   = "Monthly Revenue (INR)"
chart.y_axis.title = "Revenue"
chart.x_axis.title = "Month"
chart.style   = 10
chart.width   = 20
chart.height  = 12
chart.grouping = "clustered"
 
data = Reference(ws2, min_col=2, min_row=3, max_row=3 + n_months)
cats = Reference(ws2, min_col=1, min_row=4, max_row=3 + n_months)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.series[0].graphicalProperties.solidFill = DARK_TEAL
 
ws2.add_chart(chart, f"H3")
 
 
wb.save(output_file)
print(f"   ✅ Formatting applied")
 
print()
print("=" * 52)
print(f"  ✅  Report saved → {output_file}")
print(f"  📊  Sheets: Executive Summary | Monthly Trend |")
print(f"             Category | Top States | Fulfilment |")
print(f"             Cancellations")
print("=" * 52)

